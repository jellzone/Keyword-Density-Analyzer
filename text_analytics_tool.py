import nltk
from nltk.util import ngrams
from collections import Counter
from openpyxl import load_workbook, Workbook
import textstat
from collections import defaultdict
nltk.download("vader_lexicon")
nltk.download("punkt")
nltk.download("averaged_perceptron_tagger")

def get_ngrams(text, n):
    tokens = nltk.word_tokenize(text.lower())
    return list(ngrams(tokens, n))

def get_data_from_file(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=1, max_col=5, values_only=True):
        data.extend(list(row))
    return data

import re

def keyword_density(data):
    ngram_counts = defaultdict(int)
    for sentence in data:
        if not sentence:
            continue
        tokens = nltk.word_tokenize(sentence)
        # Remove special characters from tokens
        tokens = [re.sub(r"[^\w\s]", "", token) for token in tokens]
        tokens = [token for token in tokens if token]

        for n in range(1, 7):
            ngrams = list(nltk.ngrams(tokens, n))
            for ngram in ngrams:
                ngram_counts[ngram] += 1

    return ngram_counts

def content_analysis(data):
    full_text = " ".join([sentence for sentence in data if sentence])
    tokens = nltk.word_tokenize(full_text)
    unique_tokens = set(tokens)
    lexical_density = len(unique_tokens) / len(tokens) * 100
    return {
        "Word Count": textstat.lexicon_count(full_text),
        "Character Length": len(full_text),
        "Letters": sum(c.isalpha() for c in full_text),
        "Sentences": textstat.sentence_count(full_text),
        "Syllables": textstat.syllable_count(full_text),
        "Average Words/Sentence": textstat.avg_sentence_length(full_text),
        "Average Syllables/Word": textstat.avg_syllables_per_word(full_text),
        "Lexical Density": lexical_density,
        "Lexical Diversity": textstat.text_standard(full_text, float_output=True),
        "Reading Ease": textstat.flesch_reading_ease(full_text),
        "Grade Level": textstat.flesch_kincaid_grade(full_text),
        "Gunning Fog": textstat.gunning_fog(full_text),
        "Coleman Liau Index": textstat.coleman_liau_index(full_text),
        "Smog Index": textstat.smog_index(full_text),
        "Automated Reading Index": textstat.automated_readability_index(full_text),
    }

from nltk.sentiment import SentimentIntensityAnalyzer

def sentence_analysis(data):
    sia = SentimentIntensityAnalyzer()
    sentence_data = []
    for sentence in data:
        if not sentence:
            continue
        sentiment = sia.polarity_scores(sentence)
        readability_scores = {
            "Reading Ease": textstat.flesch_reading_ease(sentence),
            "Grade Level": textstat.flesch_kincaid_grade(sentence),
            "Gunning Fog": textstat.gunning_fog(sentence),
            "Coleman Liau Index": textstat.coleman_liau_index(sentence),
            "Smog Index": textstat.smog_index(sentence),
            "Automated Reading Index": textstat.automated_readability_index(sentence),
        }
        sentence_data.append({
            "Sentence": sentence,
            "Compound": sentiment["compound"],
            "Words": textstat.lexicon_count(sentence),
            **readability_scores
        })
    return sentence_data


def write_to_excel(data, content_data, sentence_data, output_file):
    wb = Workbook()
    keyword_density_ws = wb.active
    keyword_density_ws.title = "Keyword density"
    keyword_density_ws.append(["Phrase Length", "Keyword Phrase", "Ngram Density"])
    for ngram, count in data.items():
        keyword_density_ws.append([len(ngram), " ".join(ngram), count])

    content_analysis_ws = wb.create_sheet("Content Analysis")
    for key, value in content_data.items():
        if key == "Readability":
            for sub_key, sub_value in value.items():
                content_analysis_ws.append([sub_key, sub_value])
        else:
            content_analysis_ws.append([key, value])

    sentence_analysis_ws = wb.create_sheet("Sentence Analysis")
    sentence_analysis_ws.append(["Sentence", "Compound", "Words", "Reading Ease", "Grade Level", "Gunning Fog", "Coleman Liau Index", "Smog Index", "Automated Reading Index"])
    for s_data in sentence_data:
        sentence_analysis_ws.append([s_data["Sentence"], s_data["Compound"], s_data["Words"], s_data["Reading Ease"], s_data["Grade Level"], s_data["Gunning Fog"], s_data["Coleman Liau Index"], s_data["Smog Index"], s_data["Automated Reading Index"]])

    wb.save(output_file)

def main():
    input_file = "your_file.xlsx"
    output_file = "data_analytics.xlsx"

    data = get_data_from_file(input_file)
    ngram_counts = keyword_density(data)
    content_data = content_analysis(data)
    sentence_data = sentence_analysis(data)

    write_to_excel(ngram_counts, content_data, sentence_data, output_file)

if __name__ == "__main__":
    main()

